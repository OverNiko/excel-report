#!/usr/bin/env python 3

import pandas as pd
from datetime import datetime, timedelta
from typing import Union, Optional
from openpyxl import load_workbook
import argparse
import json

pd.set_option('future.no_silent_downcasting', True)

DEFAULT_EXCEL_FILE = 'Черненко Александр Александрович.xlsx'
BASE_DATE = datetime(2024, 10, 3)

def save_report_json(data: dict, file_name: str) -> None:
    """Сохраняет отчет в JSON файл."""
    with open(file_name, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)
    print(f"Отчет успешно сохранен в файл: {file_name}")
    
def get_attendance_status(percentage: float) -> tuple[str, str]:
    """Возвращает статус и цвет по проценту посещаемости."""
    if percentage >= 50:
        return "большинство студентов посещает занятия", "green"
    elif percentage < 50 and percentage >= 25:
        return "большинство студентов не посещает занятия", "red"
    else:
        return "студенты не посещают занятия", "red"
    
def get_student_attendance_status(percentage: float) -> tuple[str, str]:
    """Статус посещаемости для одного студента."""
    if percentage >= 75:
        return "студент хорошо посещает занятия", "green"
    elif percentage >= 50:
        return "студент удовлетворительно посещает занятия", "orange"
    elif percentage >= 25:
        return "студент редко посещает занятия", "red"
    else:
        return "студент практически не посещает занятия", "red"

def get_sheet_names(excel_file: str) -> list:
    """Получает список имен всех листов в Excel файле."""
    workbook = load_workbook(excel_file, read_only=True)
    return workbook.sheetnames

def choose_sheet(excel_file: str, sheet_index: int) -> str:
    """Позволяет пользователю выбрать лист для работы."""
    sheet_names = get_sheet_names(excel_file)
    if 0 <= sheet_index < len(sheet_names):
        return sheet_names[sheet_index]
    else:
        raise ValueError("Неверный номер листа.")

def parse_date(date: Union[datetime, str, float]) -> Optional[datetime]:
    """Преобразует различные форматы дат в объект datetime."""
    if isinstance(date, (datetime, str)):
        return pd.to_datetime(date)
    number = int(float(str(date)))
    return BASE_DATE + timedelta(days=(number - 1) * 7)

def save_report(content: str, file_name: str) -> None:
    """Сохраняет отчет в HTML файл."""
    with open(file_name, 'w', encoding='utf-8') as f:
        f.write(content)
    print(f"Отчет успешно сохранен в файл: {file_name}")

def generate_attendance_report(student_data: pd.DataFrame, dates: list, date: datetime, group_name: str, report_format: str) -> None:
    """Генерирует отчет о посещаемости за указанное занятие и сохраняет его."""
    date_str = date.strftime('%Y-%m-%d')
    if date_str in [d.strftime('%Y-%m-%d') for d in dates]:
        index = dates.index(date)
        attendance_column = student_data.columns[4 + index]
                
        attendance_data = student_data[['Фамилия', 'Имя', 'Отчество', attendance_column]].copy()
        attendance_data[attendance_column] = attendance_data[attendance_column].fillna(0).astype(int)
        
        total_students = len(student_data)
        total_attendance = sum(attendance_data[attendance_column])
        attendance_percentage = (total_attendance / total_students) * 100
        
        attendance_status, status_color = get_attendance_status(attendance_percentage)
        report_content = (
            f"<h1>Отчет по группе за дату</h1>\n"
            f"<h2>Группа: {group_name}</h2>\n"
            f"<h3>Всего студентов: {total_students}</h3>\n"
            f"<h3>Посещаемость: </h3>\n"
            f"<h4>- общая посещаемость на занятии {date.date()}: {total_attendance} из {total_students} ({attendance_percentage:.2f}%)</h4>\n"
            f"<h4>- оценка посещаемости: <span style='color:{status_color}'>{attendance_status}</span></h4>\n"
        )
        report_content += attendance_data.to_html(index=False)
        
        json_data = {
            "group": group_name,
            "date": date_str,
            "total_students": total_students,
            "attendance": int(total_attendance),
            "attendance_percentage": round(attendance_percentage, 2),
            "attendance_status": attendance_status,
            "students": attendance_data.to_dict(orient="records")
        }
        
        if report_format in ("2", "3"):
            save_report(report_content, f"отчет_группы_{group_name}_за_{date_str}.html")
        if report_format in ("1", "3"):
            save_report_json(json_data, f"отчет_группы_{group_name}_за_{date_str}.json")
    else:
        print(f"Нет занятий на дату {date_str}.")

def view_all_data(student_data: pd.DataFrame, dates: list, group_name: str, report_format: str) -> None:
    """Просмотр всех данных файла и сохраняет в отчет."""
    total_classes = len(dates)
    total_students = len(student_data)
    
    attendance_data = pd.DataFrame({
        'Дата': dates
    })
    
    for i, student in student_data.iterrows():
        attendance_data[f'Студент_{i}'] = student.iloc[4:].values
        
    attendance_data = attendance_data.fillna(0).astype(int)
    
    total_attendance = attendance_data.drop(columns=['Дата']).sum().sum()
    attendance_percentage = (total_attendance / (total_classes * total_students)) * 100
    
    attendance_status, status_color = get_attendance_status(attendance_percentage)
    report_content = (
        f"<h1>Отчет по группе</h1>\n"
        f"<h2>Группа: {group_name}</h2>\n"
        f"<h3>Всего студентов: {total_students}</h3>\n"
        f"<h3>Посещаемость: </h3>\n"
        f"<h4>- общая посещаемость на занятиях: {attendance_percentage:.2f}%</h4>\n"
        f"<h4>- оценка посещаемости: <span style='color:{status_color}'>{attendance_status}</span></h4>\n"
    )
    report_content += student_data.to_html(index=False)

    
    json_data = {
        "group": group_name,
        "total_students": total_students,
        "attendance_percentage": round(attendance_percentage, 2),
        "attendance_status": attendance_status,
        "students": student_data.to_dict(orient="records")
    }
    
    if report_format in ("2", "3"):
        save_report(report_content, f"полный_отчет_группы_{group_name}.html")
    if report_format in ("1", "3"):
        save_report_json(json_data, f"полный_отчет_группы_{group_name}.json")

def generate_student_report(student_data: pd.DataFrame, dates: list, group_name: str, student_index: int, report_format: str) -> None:
    """Генерирует отчет о посещаемости для выбранного студента и сохраняет его."""
    sorted_student_data = student_data.sort_values('№ п/п')
    if 0 <= student_index < len(sorted_student_data):
        student = sorted_student_data.iloc[student_index]
    else:
        raise ValueError("Неверный номер студента.")
    
    attendance_data = pd.DataFrame({
        'Дата': dates,
        'Присутствие': student.iloc[4:].values
    })
    attendance_data['Присутствие'] = attendance_data['Присутствие'].fillna(0).astype(int)
    
    fio = f"{student['Фамилия']} {student['Имя']} {student['Отчество']}"
    attendance_count = attendance_data['Присутствие'].sum()
    total_classes = len(dates)
    attendance_percentage = (attendance_count / total_classes) * 100
    
    attendance_status, status_color = get_student_attendance_status(attendance_percentage)
    report_content = (
        f"<h1>Отчет по студенту</h1>\n"
        f"<h2>Группа: {group_name}</h2>\n"
        f"<h3>Всего студентов: {len(student_data)}</h3>\n"
        f"<h3>ФИО: {fio}</h3>\n"
        f"<h3>Посещаемость: </h3>\n"
        f"<h3>- посещено занятий {attendance_count} / всего занятий {total_classes} -- {attendance_percentage:.2f}%</h3>\n"
        f"<h3>- пценка посещаемости: <span style='color:{status_color}'>{attendance_status}</span></h3>\n"
    )
    report_content += attendance_data.to_html(index=False)
    
    attendance_data['Дата'] = attendance_data['Дата'].astype(str)
    json_data = {
        "group": group_name,
        "fio": fio,
        "attendance_count": int(attendance_count),
        "total_classes": total_classes,
        "attendance_percentage": round(attendance_percentage, 2),
        "attendance_status": attendance_status,
        "attendance": attendance_data.to_dict(orient="records")
    }
    if report_format in ("2", "3"):
        save_report(report_content, f"отчет_о_посещаемости_студента(ки)_{student['Фамилия']}_группы_{group_name}.html")
    if report_format in ("1", "3"):
        save_report_json(json_data, f"отчет_о_посещаемости_студента(ки)_{student['Фамилия']}_группы_{group_name}.json")

def main(excel_file: str, sheet_index: int, action: str, student_index: Optional[int] = None, date_index: Optional[int] = None, report_format: str = "1") -> None:
    """Основная функция программы для обработки данных из Excel и взаимодействия с пользователем."""   
    sheet_name = choose_sheet(excel_file, sheet_index)
    
    try:
        df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
    except Exception as e:
        print(f"Ошибка при чтении файла Excel: {e}")
        return

    group_name = df.iloc[0, 0].split(':')[-1].strip()
    dates = df.iloc[2, 4:].dropna().apply(parse_date).tolist()
    # Оставляем только даты, которые уже наступили (или сегодня)
    today = datetime.now()
    dates = [d for d in dates if d <= today]
    headers = ['№ п/п', 'Фамилия', 'Имя', 'Отчество'] + [f'Занятие_{i+1}' for i in range(len(dates))]
    student_data = df.iloc[3:].reset_index(drop=True)
    student_data = student_data.iloc[:, :4+len(dates)]  # Обрезаем лишние столбцы, если даты были отброшены
    student_data.columns = headers
    student_data.iloc[:, 4:] = student_data.iloc[:, 4:].fillna(0).astype(int)

    if action == '1':
        if student_index is not None:
            generate_student_report(student_data, dates, group_name, student_index, report_format)
        else:
            print("Необходимо указать номер студента для отчета.")
    elif action == '2':
        if date_index is not None:
            if 0 <= date_index < len(dates):
                generate_attendance_report(student_data, dates, dates[date_index], group_name, report_format)
            else:
                print("Неверный номер занятия.")
        else:
            print("Необходимо указать номер занятия для отчета.")
    elif action == '3':
        view_all_data(student_data, dates, group_name, report_format)
    else:
        print("Неверный выбор. Пожалуйста, попробуйте снова.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Программа для работы с данными о посещаемости студентов.")
    parser.add_argument("--file", default=DEFAULT_EXCEL_FILE, help="Путь к файлу Excel")
    parser.add_argument("--sheet_index", default=0, type=int, help="Номер листа для работы")
    parser.add_argument("--action", default="3", help="Действие: 1 - отчет по студентам, 2 - отчет по дате, 3 - полный отчет (по умолчанию)")
    parser.add_argument("--student_index", type=int, help="Номер студента для отчета (только для действия 1)")
    parser.add_argument("--date_index", type=int, help="Номер занятия для отчета (только для действия 2)")
    parser.add_argument("--report_format", default="1", help="Формат отчета: 1 - json (по умолчанию), 2 - html, 3 - json и html")
    args = parser.parse_args()

    main(args.file, args.sheet_index, args.action, args.student_index, args.date_index, args.report_format)